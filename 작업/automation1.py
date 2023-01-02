# -*- coding: utf-8 -*-
import time
import pyautogui
import pyperclip
import win32com.client
from random import randint
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from openpyxl import load_workbook

load_wb = load_workbook("C:/Users/DW-PC/Desktop/이대명/DESK/공수/프로젝트 이력조회_20220922091029.xlsx", data_only=True)
load_ws = load_wb['Sheet1']
load_wb.create_sheet(title = "수정한이력")
# print(load_ws['A4'].value)

date_list = []
count_date = 0
work_list_time = []
work_list = []

print(len(load_ws['O']))

for x in range(4, len(load_ws['O']) + 1):
    work_type = load_ws.cell(row=x, column=14).value
    work_id = load_ws.cell(row=x, column=1).value
    tmp = load_ws.cell(row=x, column=15).value
    if(work_type == "종료" and tmp[6] == '9'):
        work_list_time.append(tmp)
        work_list.append(work_id)
    else:
        continue

print(len(work_list))
print(len(work_list_time))
print(work_list)
print(work_list_time)

    # if "-" in str(x.value):
        # date_list.append()

# for cell in load_ws['O']:
    # print(cell.value)
    # if "PM" in str(cell.value):
    #     work_list.append(cell.value)

# print(len(work_list))
# print(work_list)

# 승인일시의 월에 '9'를 포함한 것만 work_list에 포함시키기

# # Admin 사이트 접속
# driver = webdriver.Chrome("chromedriver") # Webdriver 파일의 경로를 입력
# driver.get('http://172.20.16.43/xefc/egene/login_dw.jsp') # 이동을 원하는 페이지 주소 입력
# driver.implicitly_wait(15) # 페이지 다 뜰 때 까지 기다림

# # 로그인
# driver.find_element(By.XPATH,"/html/body/form/input[1]").click()
# pyautogui.write("admin")
# driver.find_element(By.XPATH,"//input[@value='로그인']").click()
# driver.implicitly_wait(5)

# driver.get('http://172.20.16.43/#MEN00696/_14c02e795e155')






# driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div/div[2]/div[2]/div[2]/div/div/div/div/div/form/div/div/div/div/div[1]/div/div[2]/div[2]/div/input[3]").click()
# pyautogui.hotkey("ctrl", "a")
# pyautogui.press("delete")
# pyperclip.copy("2022-08-01")
# pyautogui.hotkey('ctrl', 'v')

# driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div/div[2]/div[2]/div[2]/div/div/div/div/div/form/div/div/div/div/div[1]/div/div[2]/div[2]/div/input[4]").click()
# pyautogui.hotkey("ctrl", "a")
# pyautogui.press("delete")
# pyperclip.copy("2022-08-31")
# pyautogui.hotkey('ctrl', 'v')